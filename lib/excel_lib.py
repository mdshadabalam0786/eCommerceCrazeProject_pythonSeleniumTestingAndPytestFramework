import openpyxl
def loginLocator():
    li=[]
    filePath=r'D:\scapper\ShoppersStackProject1\data\Locator.xlsx'
    wb=openpyxl.load_workbook(filePath)
    ws=wb['locatorLogin']
    result=ws.iter_rows(values_only=True)
    header=next(result)
    for i in result:
        if  i[0]!=None:
            li.append(i)

    return li #[('xpath', '//button[@id="loginBtn"]'), ('xpath', '//input[@id="Email"]'), ('xpath', '//input[@id="Password"]'), ('xpath', '//span[text()="Login"]')]




    


def registerLocator():
    regList=[]
    filepath=r"D:\scapper\ShoppersStackProject1\data\Locator.xlsx"
    wb=openpyxl.load_workbook(filepath)
    ws=wb['locatorRegister']
    result=ws.iter_rows(values_only=True)
    header=next(result)
    for i in result:
        regList.append((i[0],i[1]))
    return regList


def getDataFromExcelValidLoginCredential():
  filePath=r"D:\scapper\ShoppersStackProject1\data\TestData.xlsx"
  wb=openpyxl.load_workbook(filePath)
  ws=wb['Login']
  result=ws.iter_rows(values_only=True)
  return result

def registerData():
    regData=[]
    filePath=r"D:\scapper\ShoppersStackProject1\data\TestData.xlsx"
    ws=openpyxl.load_workbook(filePath)
    wb=ws['Register']
    result=wb.iter_rows(values_only=True)
    header=next(result)
    for i in result:
        regData.append(i)
    return regData

    

